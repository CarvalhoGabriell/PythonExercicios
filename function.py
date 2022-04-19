from openpyxl import Workbook, load_workbook
# wb = Workbook()

# sheet = wb.active

# sheet["A1"] = "Hello"
# sheet["B1"] = "Word"

# wb.save(filename="./test.xlsx")


wb = load_workbook("Criptos.xlsx")
aba_active = wb.active
print(wb)
letras = ['A1','B1','C1','D1','E1']
for col in letras:
    if col == 'A1':
        aba_active[col] = "Produto test"
    elif col == 'B1':
        aba_active[col] = "23121999"
    elif col == 'C1':
        aba_active[col] = "Assistencial"
    elif col == 'D1':
        aba_active[col] = 23.00
    else:
        aba_active[col] = 250.00


wb.save('Criptos2.xlsx')
    